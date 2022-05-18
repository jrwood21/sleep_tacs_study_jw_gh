"""
Title: Paired associate word learning task [replication of Marshall et al. 2006]
Author: Julia Wood, the University of Queensland, Australia
Code adapted from Tom Hardwicke's finger tapping task code: https://github.com/TomHardwicke/finger-tapping-task
Developed in Psychopy v2022.1.1
See my GitHub for further details: https://github.com/jrwood21
"""

# import useful modules
import time
import pandas as pd
import numpy as np
import os
import openpyxl
import random
import psychtoolbox
from psychopy import prefs, visual, event, core, gui
prefs.hardware['audioLib'] = 'PTB' # change the audio library to psychtoolbox for best latencies
prefs.hardware['audioLatencyMode'] = 3 # set the latency mode to high precision 
prefs.hardware['audioDriver'] = 'Primary Sound'
from psychopy import sound # must import sound after changing sound prefs above

micDevice = sound.Microphone.getDevices()[0] # define mic device explicitly

os.chdir(os.path.abspath(''))  # change working directory to script directory
globalClock = core.Clock()  # create timer to track the time since experiment started
mic=sound.Microphone(channels=1, streamBufferSecs=10, device=micDevice) # buffersecs is the length of the mic recording. Use mic.poll() below to extend recording time

### set up some useful functions ###
# Function to save messages to a log file recording everything the exp is doing
def saveToLog(logString, timeStamp=1):
    f = open(logFile, 'a')  # open our log file in append mode 
    f.write(logString)  # write the text
    if timeStamp != 0:  # if timestamp has not been turned off
        f.write('// logged at %iseconds' % globalClock.getTime())  # write a timestamp (very coarse)
    f.write('\n')  # create new line
    f.close()  # close and save the log file

# An exit function to initiate if end key is pressed
def quitExp():
    if 'logFile' in globals():  # if a log file has been created
        saveToLog('User aborted experiment')
        saveToLog('..........................................', 0)
    if 'win' in globals():  # if a window has been created
        win.close()  # close the window
    core.quit()  # quit the program

# define function to check if a filename exists. If yes, iterate to get the next available file number (avoid overwriting)
def uniq_path(path):
    fn, ext = os.path.splitext(path)
    counter = 2
    while os.path.exists(path):
        path = fn + "_" + str(counter) + ext
        counter += 1
    return path

# function to convert workbooks to ordered word lists
def asWordLists(sheet, n_items):
    cue_words = [] # store cue words and matching recall words in lists
    recall_words = []
    stim_order = [] # create a counter to use for randomisation of wordpair presentation order
    ix = 0 # create counter
    for rowx in sheet.iter_rows(max_row=n_items): 
        cue_words.append(rowx[0].value) # read cue word
        recall_words.append(rowx[1].value) # read in matching recall word
        stim_order.append(ix) # generate list of numbers the length of number of rows
        ix += 1
    return cue_words, recall_words, stim_order

# Function to generate random word lists from ordered word lists
def randomWordLists(cue_words, recall_words, stim_order, num_items):
    random.seed()
    random.shuffle(stim_order) # shuffle number list to randomise order of presentation
    cue_words_2 = [] # create randomised cue and recall word lists to use in experiment (but pairs remain matched)
    recall_words_2 = []
    for i in range(num_items): 
        cue_words_2.append(cue_words[stim_order[i]]) # generate cue words in randomised order
        recall_words_2.append(recall_words[stim_order[i]]) # generate recall words in randomised order
    return cue_words_2, recall_words_2

# Function to display wordlists
def displayWordListPairs(num_words, cue_wordlist, recall_wordlist):
    for i in range(num_words):
        if event.getKeys(['end']):  # checks for quit routine
            quitExp()  
        cueWordListText.setText(cue_wordlist[i]) # iterate over cue words and set the text
        recallWordListText.setText(recall_wordlist[i]) # iterate over matching recall words and set the text
        cueWordListText.setAutoDraw(False) # set autodraw to false, otherwise text will display in ALL frames
        recallWordListText.setAutoDraw(False)
        cueWordListText.draw()
        recallWordListText.draw()
        win.flip() # display the text
        core.wait(5) # show text for 5 seconds
        win.flip() # blank the screen
        core.wait(0.1) # wait 100ms

# Function to run word learning phase of task
def wordLearning(wordlist, wordlist_type, workbook="wordlists_audio.xlsx"):
    ## Intro screen ##
    saveToLog('Presenting word learning task introduction screen') # save info to log
    win.setColor('#000000', colorSpace='hex')  # set background colour to black
    win.flip()  # blank the screen first
    generalText.setText(
        'TASK INSTRUCTIONS \n\nOver the next 5 minutes, you will be shown a slide show of word pairs. Each word pair will be displayed for 5 seconds. The words in each pair are related to each other. \nFor example, PLANET and MARS. \n\nPlease try to memorise each of the word pairs as best you can, as you will be asked to recall these word pairs in a later task. \n\nTo help you memorise each word pair, please associate each of the word pairs with a story. \nFor example, MARS is the fourth PLANET from the Sun. \n\nPress the spacebar when you are ready to commence the task.')
    generalText.draw() # draw the text
    win.flip()  # show the text in the window
    event.waitKeys(keyList=["space"])  # wait for a spacebar press before continuing
    if ((task_attempt_number > 1) and (event.getKeys(['end']))): # include option to save data on attempts >1, in case use accidentally selected option to re-enter loop
        res['pc30_trial_num'] = task_attempt_number - 1
        res.to_csv(fileName)
        saveToLog('User quit the experiment before entering another loop of learning tasks', 0)
        quitExp()
    event.clearEvents()  # clear the event buffer
    win.flip()  # blank the screen first
    core.wait(2)
    saveToLog('Running word learning task with wordlist %s' % (wordlist_type))  # save info to log
    
    wordlist_book = openpyxl.load_workbook(workbook) # read in word lists from xlsx document
    if wordlist == 'wordlist_1':
        wordlist_sheet = wordlist_book.worksheets[0] # specify which xlsx sheet to use
    elif wordlist == 'wordlist_2':
        wordlist_sheet = wordlist_book.worksheets[1]
    elif wordlist == 'wordlist_prac':
        wordlist_sheet = wordlist_book.worksheets[2]
    
    ### create word lists and randomise them (pairwise)
    c_words, r_words, s_order = asWordLists(sheet=wordlist_sheet, n_items=46)
    rand_c_words, rand_r_words = randomWordLists(cue_words=c_words, recall_words=r_words, stim_order=s_order, num_items=46)
    
    #### import dummy word lists and convert to lists, then randomise order (pairwise)
    if wordlist == 'wordlist_1':
        dummy_wordlist1_sheet = wordlist_book.worksheets[3]
    elif wordlist == 'wordlist_2':
        dummy_wordlist1_sheet = wordlist_book.worksheets[4]
    elif wordlist == 'wordlist_prac': # prac word list defined as a dummy word list
        dummy_wordlist1_sheet = wordlist_book.worksheets[2]
    dummy_c_words, dummy_r_words, dummy_s_order = asWordLists(sheet=dummy_wordlist1_sheet, n_items=8)
    rdum_c_words, rdum_r_words = randomWordLists(cue_words=dummy_c_words, recall_words=dummy_r_words, stim_order=dummy_s_order, num_items=8)
    rand_dummy_c_words = rdum_c_words[0:4] # only select the first 4
    rand_dummy_r_words = rdum_r_words[0:4]

    win.setColor('#000000', colorSpace='hex') # set the background colour to black and clear the screen
    win.flip() 
    
    # display 4 x dummy word pairs in random order for 5s each with 100ms ISI
    displayWordListPairs(num_words=4, cue_wordlist=rand_dummy_c_words, recall_wordlist=rand_dummy_r_words)
    
    if not metaData['practice mode']: # if it is not practice mode, display the test word pair list in random order
        displayWordListPairs(num_words=46, cue_wordlist=rand_c_words, recall_wordlist=rand_r_words) 
    
    rand_dummy_c_words = rdum_c_words[4:8] # now select the last 4 dummy word pairs from randomised list
    rand_dummy_r_words = rdum_r_words[4:8]
    
    # display 4 x dummy word pairs in random order for 5s each with 100ms ISI
    displayWordListPairs(num_words=4, cue_wordlist=rand_dummy_c_words, recall_wordlist=rand_dummy_r_words)
    
    # gather all relevant data for this trial in a dictionary
    if not metaData['practice mode']:  
        newRow = {'participant': metaData['participant'], 
                  'participant_allocation': metaData['participant allocation'],
                  'session': metaData['session number'],
                  'session_time': metaData['session time'],
                  'task_type': 'learning',
                  'wordlist_type': wordlist_type}
    
    elif metaData['practice mode']: 
        newRow = {'participant': metaData['participant'], 
                  'session': metaData['session number'],
                  'session_time': metaData['session time'],
                  'task_type': 'learning',
                  'wordlist_type': wordlist_type}
    
    # export word pair presentation order to csv
    pair_order = np.arange(1,47,1)
    pres_order_df = pd.DataFrame({'presentation_order': pair_order,
                                  'cue_word': rand_c_words,
                                  'response_word': rand_r_words})
    pres_path = p_dir + os.path.sep + 'P' + str(metaData['participant']) + "_" + str(metaData['participant allocation']) + '_S' + str(metaData['session number']) + '_' + str(metaData['session time']) + '_LEARNING_WORD_ORDER.csv'
    pres_path = uniq_path(pres_path)
    pres_order_df.to_csv(pres_path)
    saveToLog('Learning word presentation order saved to %s' % (pres_path))
    
    # convert session data to df
    store_out = pd.DataFrame(newRow, index=[0])

    # record wordlist type used in metadata log file:
    metaData.update({'wordlist type': wordlist_type})
    saveToLog('learning task completed with wordlist %s' % (metaData['wordlist type']))
    win.flip()  # blank the screen
     
    return store_out
    
# Function to execute cued recall task only
def wordRecall(wordlist, wordlist_type, workbook="wordlists_audio.xlsx"):
    ## Intro screen ##
    saveToLog('Presenting word recall task introduction screen') # save info to log
    win.setColor('#000000', colorSpace='hex')  # set background colour to black
    win.flip()  # blank the screen first
    generalText.setText(
        'TASK INSTRUCTIONS\n\nOne word from each of the word pairs you viewed earlier will now be presented to you on the computer screen. Only a single word will be shown at once. \n\nFor each word presented to you, please try to recall the matching word from its pair. Say your answer out loud.  \n\nIf you cannot recall the matching word, please advise the experimenter that you would like to skip to the next word.  \n\nPress the spacebar when you are ready to commence the task.')
    generalText.draw() # draw the text
    win.flip()  # show the text in the window
    event.waitKeys(keyList=["space"])  # wait for a spacebar press before continuing
    event.clearEvents()  # clear the event buffer
    win.flip()  # blank the screen first
    core.wait(2)
    saveToLog('Running word recall task with wordlist %s' % (wordlist_type))  # save info to log
    
    # check that the file directory for this participant's audio files is set up:
    audio_dir = p_dir + os.path.sep + 'audio_recall_files'
    if not os.path.isdir(audio_dir):
        os.mkdir(audio_dir)
    audio_ses_dir = audio_dir + os.path.sep + 'S' + str(metaData['session number']) + '_' + str(metaData['session time'])
    if not os.path.isdir(audio_ses_dir):
        os.mkdir(audio_ses_dir)
    
    wordlist_book = openpyxl.load_workbook(workbook) # read in word lists from xlsx document
    if wordlist == 'wordlist_1':
        wordlist_sheet = wordlist_book.worksheets[0] # specify which sheet to use
        n_words = 46
    elif wordlist == 'wordlist_2':
        wordlist_sheet = wordlist_book.worksheets[1]
        n_words = 46
    elif wordlist == 'wordlist_prac':
        wordlist_sheet = wordlist_book.worksheets[2]
        n_words = 8
    
    ### create word lists and randomise (pairwise)
    c_words, r_words, s_order = asWordLists(sheet=wordlist_sheet, n_items=n_words)
    rand_c_words, rand_r_words = randomWordLists(cue_words=c_words, recall_words=r_words, stim_order=s_order, num_items=n_words)
    
    win.setColor('#000000', colorSpace='hex') # set the background colour to black and clear the screen
    win.flip() 
    myMouse = event.Mouse() # define mouse
    myMouse.clickReset() # clear all mouse click events
    event.clearEvents()  # clear the event buffer (all events)
    
    # create lists to store word pair presentation order and timings
    ## NOTE THAT THESE TIMINGS ARE NOT RESPONSE TIMES - AUDIO FILES MUST BE ANALYSED IN SEPARATE SOFTWARE TO GET RESPONSE TIME!!!
    # These timings provide coarse estimates of latencies in the code execution
    order = []
    cue_word = []
    response_word = []
    recall_loop_start_times = []
    text_draw_times = []
    cue_word_times = []
    mic_start_times = []
    mic_stop_times = []
    
    # display each cue word on it's own (random order), then display matching recall word after a mouse click
    for i in range (n_words):
        order.append(i+1)
        cueWordListText_recall.setText(rand_c_words[i]) # set the cue word
        cue_word.append(rand_c_words[i])
        recallWordListText_recall.setText(rand_r_words[i]) # set the recall word
        response_word.append(rand_r_words[i])
        cueWordListText_recall.setAutoDraw(False) # set autodraw to false, otherwise text will display in ALL frames
        recallWordListText_recall.setAutoDraw(False)
        
        recall_loop_start_times.append(time.perf_counter_ns()/1000000) # get system time in milliseconds
        cueWordListText_recall.draw()
        text_draw_times.append(time.perf_counter_ns()/1000000)
        win.flip() # display the cue word
        cue_word_times.append(time.perf_counter_ns()/1000000)
        mic.start()
        mic_start_times.append(time.perf_counter_ns()/1000000)
        
        buttons = myMouse.getPressed(getTime=False) # check for mouse clicks
        while buttons == [0,0,0]: # while there are no mouse clicks
            buttons = myMouse.getPressed(getTime=False) # keep checking for mouse clicks
            mic.poll()
            if buttons != [0,0,0]: # when a mouse click is registered
                mic.stop()
                mic_stop_times.append(time.perf_counter_ns()/1000000)
                audioclip = mic.getRecording()
                audio_path = audio_ses_dir + os.path.sep + 'P' + str(metaData['participant']) + "_" + str(metaData['participant allocation']) + '_S' + str(metaData['session number']) + '_' + str(metaData['session time']) + '_' + str(rand_c_words[i]) + '_' + str(rand_r_words[i]) + '.wav'
                audio_path = uniq_path(audio_path)
                audioclip.save(audio_path)
                break # exit the loop
            if event.getKeys(['end']):  # if the user hits the 'end' key
                mic_stop_times.append(np.nan)
                lists_to_df = pd.DataFrame({'order': order, # export word pair presentation order so far
                            'cue_word': cue_word,
                            'response_word': response_word,
                            'recall_loop_start_time': recall_loop_start_times,
                            'text_draw_times': text_draw_times,
                            'cue_word_times': cue_word_times,
                            'mic_start_times': mic_start_times,
                            'mic_stop_times': mic_stop_times})
                list_path = audio_ses_dir + os.path.sep + 'P' + str(metaData['participant']) + "_" + str(metaData['participant allocation']) + '_S' + str(metaData['session number']) + '_' + str(metaData['session time']) + '_RECALL_WORD_ORDER_quitExp.csv'
                list_path = uniq_path(list_path)
                lists_to_df.to_csv(list_path)
                saveToLog('User quit the experiment. In-progress recall word presentation order and time data saved to %s' % (list_path))
                quitExp()  # quit the experiment
        win.flip() # blank the screen
        core.wait(0.1)
        event.clearEvents()
        
        if metaData['session time'] == 'pm-a': # if it is the learning phase of the pm session
            recallWordListText_recall.draw() # provide accuracy feedback after each cue word
            win.flip() 
            core.wait(2.5) # display accuracy feedback for 2.5sec
            win.flip()
            core.wait(0.1)
            event.clearEvents()
    
    lists_to_df = pd.DataFrame({'order': order, # export recall word presentation to csv
                                'cue_word': cue_word,
                                'response_word': response_word,
                                'recall_loop_start_time': recall_loop_start_times,
                                'text_draw_times': text_draw_times,
                                'cue_word_times': cue_word_times,
                                'mic_start_times': mic_start_times,
                                'mic_stop_times': mic_stop_times})
    list_path = audio_ses_dir + os.path.sep + 'P' + str(metaData['participant']) + "_" + str(metaData['participant allocation']) + '_S' + str(metaData['session number']) + '_' + str(metaData['session time']) + '_RECALL_WORD_ORDER.csv'
    list_path = uniq_path(list_path)
    lists_to_df.to_csv(list_path)
    saveToLog('Recall word presentation order and time data saved to %s' % (list_path))

    # gather all relevant data for this trial in a dictionary
    newRow = {'participant': metaData['participant'], 
              'participant_allocation': metaData['participant allocation'],
              'session': metaData['session number'],
              'session_time': metaData['session time'],
              'task_type': 'recall',
              'wordlist_type': wordlist_type}

    # convert session data to df
    store_out = pd.DataFrame(newRow, index=[0])

    # record wordlist type in metadata log file:
    metaData.update({'wordlist type': wordlist_type})
    saveToLog('Recall task completed with wordlist %s' % (metaData['wordlist type']))

    win.flip()  # blank the screen
    
    return store_out

### Collect and store metadata about the experiment session ###
expName = 'Paired associate word learning task'  # define experiment name
date = time.strftime("%d %b %Y %H:%M:%S", time.localtime())  # get date and time
metaData = {'participant': '',
            'session number': [1, 2],
            'session time': ['pm-a', 'pm-b', 'am'],
            'practice mode': False,
            'use automated counter-balancing': True,
            'researcher': 'JW',
            'location': '304, Seddon North, UQ, Brisbane'}  # set up info for infoBox gui
infoBox = gui.DlgFromDict(dictionary=metaData,
                          title=expName,
                          order=['participant', 'session number', 'session time',
                                 'practice mode','use automated counter-balancing'])  # display gui to get info from user
if not infoBox.OK:  # if user hit cancel
    quitExp()  # quit

# check if participant dir exists, and if not, create one:
p_dir = 'data' + os.path.sep + 'wordlearning' + os.path.sep + 'P' + str(metaData['participant'])
if not os.path.isdir(p_dir):
    os.mkdir(p_dir)

if not metaData['practice mode']:  # if this is not practice mode:
    if metaData['use automated counter-balancing']:  # AND the user has chosen to use automated counter-balancing
        cb = {'participant allocation': ['AJX', 'AJY', 'AKX', 'AKY',
                                         'BJX', 'BJY', 'BKX', 'BKY']}  # set up info for infoBox gui
        infoBox = gui.DlgFromDict(dictionary=cb,
                                  title='Choose counter-balancing parameters')  # display gui to get counterbalancing info from user
        metaData.update({'participant allocation': cb['participant allocation']})
        if not infoBox.OK:  # if user hit cancel
            quitExp()  # quit
    
    elif not metaData['use automated counter-balancing']: # OR if the user will manually select task and word list
        wl_dict = {'use word list number': ['wordlist_1', 'wordlist_2'],
                   'use task type': ['word learning', 'word recall']}
        infoBox = gui.DlgFromDict(dictionary=wl_dict,
                                  title='Select wordlist and task type to run experiment')  # display gui to get task and wordlist selection from user
        metaData.update({'participant allocation': 'manual_selection',
                         'wordlist type': '%s' % wl_dict['use word list number'],
                         'task type': '%s' % wl_dict['use task type']})
        if not infoBox.OK:  # if user hit cancel
            quitExp()  # quit
    
    # build filename for this participant's data
    fileName = p_dir + os.path.sep + 'P' + str(metaData['participant']) + "_" + str(metaData['participant allocation']) + '_S' + str(metaData['session number']) + '_' + str(metaData['session time']) + '.csv'

    if os.path.exists(fileName):  # check user knows sessions already exist for this participant's current session and time:
        myDlg = gui.Dlg()
        myDlg.addText( # inform user that files will be stored under a different name
            "This participant has existing files for this session time in the directory! Click ok to continue or cancel to abort. \n\n NOTE: if you choose to continue, files will be stored under a different file name.")
        myDlg.show()  # show dialog and wait for OK or Cancel
        if not myDlg.OK:  # if the user pressed cancel
            quitExp()
        fileName = uniq_path(fileName) # redefine file name by appending a number to prevent overwriting

    metaData.update({'expName': expName, 'date': date})  # record the info in the metaData
    
    # check if logfile exists for this participant. If not, create one:
    logFile = p_dir + os.path.sep + 'P' + str(metaData['participant']) + "_" + str(metaData['participant allocation']) +'_log.txt'
    if not os.path.exists(logFile):
        with open(logFile, 'w') as fp:
            pass

    # save metaData to log
    saveToLog('..........................................', 0)
    saveToLog('experiment: %s' % (metaData['expName']), 0)
    saveToLog('researcher: %s' % (metaData['researcher']), 0)
    saveToLog('location: %s' % (metaData['location']), 0)
    saveToLog('date: %s' % (metaData['date']), 0)
    saveToLog('participant: %s' % (metaData['participant']), 0)
    saveToLog('session: %s' % (metaData['session number']), 0)
    saveToLog('session time: %s' % (metaData['session time']), 0)
    saveToLog('participant allocation: %s' % (metaData['participant allocation']), 0)

else:  # if it is practice mode, set up practice log file & ask user to select task type
    logFile = p_dir + os.path.sep + 'P' + str(metaData['participant']) + '_practice_log.txt'
    if not os.path.exists(logFile):
        with open(logFile, 'w') as fp:
            pass

    prac_dict = {'use task type': ['word learning', 'word recall']}
    infoBox = gui.DlgFromDict(dictionary=prac_dict,
                              title='Select task type to run experiment')  # display gui to get info from user
    metaData.update({'participant allocation': 'practice session',
                     'wordlist type': 'wordlist_prac',
                     'task type': '%s' % prac_dict['use task type']})
    if not infoBox.OK:  # if user hit cancel
        quitExp()  # quit
    
    # save metaData to log
    saveToLog('..........................................', 0)
    saveToLog('experiment: %s' % (expName), 0)
    saveToLog('researcher: %s' % (metaData['researcher']), 0)
    saveToLog('location: %s' % (metaData['location']), 0)
    saveToLog('date: %s' % (date), 0)
    saveToLog('participant: %s' % (metaData['participant']), 0)
    saveToLog('session: %s' % (metaData['session number']), 0)
    saveToLog('session time: %s' % (metaData['session time']), 0)
    saveToLog('participant allocation: %s' % (metaData['participant allocation']), 0)
    
### Prepare stimuli etc ###
win = visual.Window(size=(1920, 1080), fullscr=False, screen=0, allowGUI=False, allowStencil=False, ### CHANGE SCREEN SIZE TO MATCH YOUR MONITOR
                    monitor='testMonitor', color=(-1,-1,-1), colorSpace='rgb', units='pix') # setup the Window
generalText = visual.TextStim(win=win, ori=0, name='generalText', text='', font=u'Arial', pos=[0, 0], height=35,
                              wrapWidth=920, color=(1,1,1), colorSpace='rgb', opacity=1, depth=0.0)  # general text settings
cueWordListText = visual.TextStim(win=win, ori=0, name='cueWordListText', text='', font=u'Arial', pos=[-250, 0], height=50,
                               wrapWidth=None, color=(1,1,1), colorSpace='rgb', opacity=1, depth=0.0)  # cue word list text settings
recallWordListText = visual.TextStim(win=win, ori=0, name='recallWordListText', text='', font=u'Arial', pos=[250, 0], height=50,
                               wrapWidth=None, color=(1,1,1), colorSpace='rgb', opacity=1, depth=0.0)  # recall word list text settings
cueWordListText_recall = visual.TextStim(win=win, ori=0, name='cueWordListText', text='', font=u'Arial', pos=[0, 0], height=50,
                               wrapWidth=None, color=(1,1,1), colorSpace='rgb', opacity=1, depth=0.0)  # cue word list text settings
recallWordListText_recall = visual.TextStim(win=win, ori=0, name='recallWordListText', text='', font=u'Arial', pos=[0, 0], height=50,
                               wrapWidth=None, color=(-1, -0.215686274509804, -1), colorSpace='rgb', opacity=1, depth=0.0)  # recall word list text settings - set text to darkgreen

saveToLog('Set up complete') # save info to log
### set-up complete ###

### run the experiment ###
recall_accuracy = 0 # set participant's accuracy score to 0 until redefined after recall task
task_attempt_number = 1 # set participant's task attempt number to 1 until redefined after multiple attempts

if metaData['practice mode']:  # if user has chosen practice mode
    if prac_dict['use task type'] == 'word learning':
        res = wordLearning(wordlist='wordlist_prac', wordlist_type='practice')  # run practice wordlist
    elif prac_dict['use task type'] == 'word recall':
        res = wordRecall(wordlist='wordlist_prac', wordlist_type='practice')

elif not metaData['practice mode']: # if it is NOT practice mode
    if not metaData['use automated counter-balancing']: # and the word list has been manually selected
        if wl_dict['use word list number'] == 'wordlist_1':
            man_wordlist_type = 'one'
        elif wl_dict['use word list number'] == 'wordlist_2':
            man_wordlist_type = 'two'

        # EITHER run LEARNING task once with word list manually selected
        if wl_dict['use task type'] == 'word learning':
            res = wordLearning(wordlist=wl_dict['use word list number'], wordlist_type=man_wordlist_type)
        
        # OR run word RECALL task with accuracy loop
        elif wl_dict['use task type'] == 'word recall':
            while recall_accuracy == 0: # while participant's score is below 30%
                res = wordRecall(wordlist=wl_dict['use word list number'], wordlist_type=man_wordlist_type) 
                # ask user if at least 30% accuracy achieved:
                myDlg = gui.Dlg(title='Recall accuracy check')
                myDlg.addText('Did the participant achieve at least 30% accuracy?')
                myDlg.addField('answer', choices=['no', 'yes'])
                myDlg.addText('NOTE: if this is a recall session only (pm-b or am), select YES to exit and record final result')
                acc_dat = myDlg.show() # show dialogue box gui
                if not myDlg.OK: # if user hit cancel
                    quitExp() # quit
                
                if acc_dat[0] == 'no': # if user selected NO, perform the recall task again
                    task_attempt_number = task_attempt_number + 1
                    recall_accuracy = 0
                elif acc_dat[0] == 'yes': # otherwise, if user selected YES
                    res['pc30_trial_num'] = task_attempt_number # save task attempt number where participant achieved >30% in csv file
                    if metaData['session time'] == 'pm-a':
                        saveToLog('At least 30 percent recall accuracy achieved on attempt number %s' % (task_attempt_number), 0) 
                    elif metaData['session time'] == 'pm-b' or 'am':
                        saveToLog('Single trial of word recall task completed with no accuracy feedback provided', 0)
                    recall_accuracy = 1
                    break

                # include option to quit, in case of looping error
                if event.getKeys(['end']): 
                    quitExp()

    # OR if automated counter balancing selected:
    elif metaData['use automated counter-balancing']: 
        while recall_accuracy == 0: # while the participant's score is less than 30%:
            
            ####### X ORDER
            if ((metaData['participant allocation'] == 'AJX') or (metaData['participant allocation'] == 'BJX') or (metaData['participant allocation'] == 'AKX') or (metaData['participant allocation'] == 'BKX')):
                # session 1
                if int(metaData['session number']) == 1:
                    if metaData['session time'] == 'pm-a':
                        res1 = wordLearning(wordlist='wordlist_1', wordlist_type='one')  # wordlist 1
                        res2 = wordRecall(wordlist='wordlist_1', wordlist_type='one') # wordlist 1
                        res = res1.append(res2, ignore_index=True)
                    elif metaData['session time'] == 'pm-b' or 'am':
                        res = wordRecall(wordlist='wordlist_1', wordlist_type='one') # wordlist 1
                # session 2
                elif int(metaData['session number']) == 2:
                    if metaData['session time'] == 'pm-a':
                        res1 = wordLearning(wordlist='wordlist_2', wordlist_type='two')  # wordlist 2
                        res2 = wordRecall(wordlist='wordlist_2', wordlist_type='two') # wordlist 2
                        res = res1.append(res2, ignore_index=True)
                    elif metaData['session time'] == 'pm-b' or 'am':
                        res = wordRecall(wordlist='wordlist_2', wordlist_type='two') # wordlist 2

            ####### Y ORDER
            elif ((metaData['participant allocation'] == 'AJY') or (metaData['participant allocation'] == 'BJY') or (metaData['participant allocation'] == 'AKY') or (metaData['participant allocation'] == 'BKY')):
                # session 1
                if int(metaData['session number']) == 1:
                    if metaData['session time'] == 'pm-a':
                        res1 = wordLearning(wordlist='wordlist_2', wordlist_type='two')  # wordlist 2
                        res2 = wordRecall(wordlist='wordlist_2', wordlist_type='two') # wordlist 2
                        res = res1.append(res2, ignore_index=True)
                    elif metaData['session time'] == 'pm-b' or 'am':
                        res = wordRecall(wordlist='wordlist_2', wordlist_type='two') # wordlist 2
                # session 2
                elif int(metaData['session number']) == 2:
                    if metaData['session time'] == 'pm-a':
                        res1 = wordLearning(wordlist='wordlist_1', wordlist_type='one')  # wordlist 1
                        res2 = wordRecall(wordlist='wordlist_1', wordlist_type='one') # wordlist 1
                        res = res1.append(res2, ignore_index=True)
                    elif metaData['session time'] == 'pm-b' or 'am':
                        res = wordRecall(wordlist='wordlist_1', wordlist_type='one') # wordlist 1

            # ask user if at least 30% accuracy achieved:
            myDlg = gui.Dlg(title='Recall accuracy check')
            myDlg.addText('Did the participant achieve at least 30% accuracy?')
            myDlg.addField('answer', choices=['no', 'yes'])
            myDlg.addText('NOTE: if this is a recall session only (pm-b or am), select YES to exit and record final result')
            acc_dat = myDlg.show() # show dialogue box gui
            if not myDlg.OK: # if user hit cancel
                quitExp() # quit

            if acc_dat[0] == 'no': # if the user selects NO, re-run the appropriate learning tasks
                task_attempt_number = task_attempt_number + 1
                recall_accuracy = 0
            elif acc_dat[0] == 'yes': # if >30% accuracy achieved, exit the loop
                res['pc30_trial_num'] = task_attempt_number # save task attempt number where participant achieved >30% in csv and logfile
                if metaData['session time'] == 'pm-a':
                    saveToLog('At least 30 percent recall accuracy achieved on attempt number %s' % (task_attempt_number), 0) 
                elif metaData['session time'] == 'pm-b' or 'am':
                    saveToLog('Single trial of word recall task completed with no accuracy feedback provided', 0)
                recall_accuracy = 1
                break

            # include option to quit, in case of looping error
            if event.getKeys(['end']): 
                quitExp()


## End screen ##
saveToLog('Presenting end screen')  # save info to log
win.setColor('#000000', colorSpace='hex')  # set background colour to black
win.flip()
generalText.setText(u'Thank you. That is the end of this section. Please inform the researcher you have finished.')
generalText.draw()
win.flip()  # present video buffer
event.waitKeys() # wait for a key press before continuing
event.clearEvents() # clear the event buffer

saveToLog('Experiment presentation over')  # save info to log
### Finished running the experiment ###

if metaData['practice mode']:  # if user has chosen practice mode
    quitExp()  # quit

### Save and clean up ###
win.close()

'''
Save the data as a csv file. If saving is not possible, this is usually because the file is already open.
Ask the user to close if this is the case.
If this does not resolve the situation, an attempt will be made to save the data with a different filename.
'''
while True:
    try:
        res.to_csv(fileName)
        saveToLog('Data saved with file name: %s' % fileName) # save info to log
        break
    except: # if cannot save data, likely because file is already open, ask user to close
        saveToLog('Problem encountered saving data - requesting user close open data files...') # save info to log
        myDlg = gui.Dlg()
        myDlg.addText("Unable to store data. Try closing open excel files and then click ok. Press cancel to attempt data storage to new file.")
        myDlg.show()  # show dialog and wait for OK or Cancel
        if not myDlg.OK:  # if the user pressed cancel
            fileName = p_dir + os.path.sep + 'P' + str(metaData['participant']) + "_ProblemSaving_" + str(metaData['participant allocation']) + '_S' + str(metaData['session number']) + '_' + str(metaData['session time']) + '.csv'
            saveToLog('Attempting to save data with different filename: %s' %fileName) # save info to log
            try:
                res.to_csv(fileName)
                print('Data was saved with a different filename: %s' %fileName)
                saveToLog('Data saved with file name: %s' % fileName) # save info to log
                break
            except:
                saveToLog('Major error: Data could not be saved') # save info to log
                quitExp() # quit the experiment

t = globalClock.getTime() # get run time of experiment
saveToLog('Total experiment runtime was %i seconds' % t) # record runtime to log
saveToLog('..........................................', 0)

# Shut down:
core.quit()



